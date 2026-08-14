"""Synthetic source-workbook builder mirroring the real sheet layouts (spec section 9).

Each default_*_rows() returns fresh row lists tests may mutate before build_workbook().
Layout quirks reproduced: two-row Net Worth header with merged-band Nones and % columns,
0.001 sentinels, 'Average' + numeric-year rollup + future template rows, zero-share
position rows, formula-error strings, time(0,0) ex-div junk, ESPP template dates and
taxation-calculator block, partial focal rows.
"""

import io
from datetime import datetime, time

import openpyxl

# Keep in sync with parse_taxes: the builder generates the 41-label input block from the
# parser's own sequence so fixture and contract cannot drift.
from app.importer.parsers import SHEET_TAX_INPUT_SEQUENCE


def default_net_worth_rows() -> list[list]:
    return [
        ["Month", "Date", "CASH", None, "PRE-TAX", None, "LIABILITIES", None, "NET WORTH", None],
        [None, None, "Checking", "%", "IRA", "%", "Credit Card", "%", None, None],
        [
            datetime(2024, 1, 1),
            datetime(2024, 1, 5),
            100.5,
            "N/A",
            0.001,
            "N/A",
            25.0,
            "N/A",
            75.5,
            "N/A",
        ],
        [datetime(2024, 2, 1), datetime(2024, 2, 1), 200.0, 0.99, 50.25, 0, 30.0, 0.2, 220.25, 1.9],
        [datetime(2024, 3, 1), datetime(2024, 3, 1), None, 0, None, 0, None, 0, None, 0],
    ]


def default_spending_rows() -> list[list]:
    return [
        [None, "Food", "Rent", "TOTAL", "Net Pay", "4% Portfolio", "Savings Rate"],
        ["Average", 10, 20, 30, None, None, None],
        [datetime(2024, 1, 1), 100.0, 900.0, 1000.0, 3000.0, 50.0, 0.5],
        [datetime(2024, 2, 1), 50.0, 900.0, 951.0, 3000.0, 50.0, 0.5],
        [2024.0, 150.0, 1800.0, 1950.0, 6000.0, None, None],
        [datetime(2024, 3, 1), None, None, None, None, None, None],
    ]


def default_positions_rows() -> list[list]:
    header = [
        "Platforms",
        "Type",
        "Stock",
        "Transacted Shares",
        "Transacted Price/ Share",
        "Fees",
        "Stock Split",
        "Prev Row",
        "Previous Shares",
        "Cumulative Shares",
        "Transacted Value",
        "Previous Cost",
    ]
    return [
        header,
        ["RH Taxable", "Buy", "Acme ETF", 10.123456789, 100.123456, None, None, 0, 0, 0, 0, 0],
        ["RH Taxable", "Buy", "Acme ETF", 0.0, 0.0, None, None, 0, 0, 0, 0, 0],
        ["Fido", "Sell", "Acme ETF", 2.0, 110.0, 1.5, None, 0, 0, 0, 0, 0],
        ["RH Taxable", "Buy", "Mystery Fund", 1.0, 25.0, None, None, 0, 0, 0, 0, 0],
    ]


def default_portfolio_rows() -> list[list]:
    header = [
        "Company Name",
        "Ticker",
        "Industry",
        "Shares",
        "Market Weight",
        "Current Price",
        "Daily Gain/Loss",
        "Daily Change %",
        "1yr Chart",
        "Cost Basis",
        "Market Value",
        "Unrealized Gain/Loss",
        "Unrealized Gain/Loss %",
        "XIRR",
        "Realized Gain/Loss",
        "Dividends Collected",
        "Total Gain/Loss",
    ]
    totals = [None, None, None, None, 1, None, 0, 0, None, 0, 0, 0, 0, 0, 0, 0, 0]
    return [
        header,
        totals,
        ["Acme ETF", "ACME", "ETF", 10, 0.5, 100.5, 0, 0, None, 0, 0, 0, 0, 0, 0, 0, 0],
        ["Div Corp", "DIVC", "Financials", 5, 0.5, 20.0, 0, 0, None, 0, 0, 0, 0, 0, 0, 12.5, 0],
    ]


def default_taxes_rows() -> list[list]:
    rows: list[list] = [["Fill in White cells", None, 2023.0, 2024.0, 2025.0]]
    # 41-label input block, values 100+i (year1) / 200+i (year2); year3 column stays empty.
    counter = 0
    for section_header, sequence in SHEET_TAX_INPUT_SEQUENCE:
        for offset, (label, key) in enumerate(sequence):
            value1, value2 = 100.0 + counter, 200.0 + counter
            if key == "unq_div_state_exempt_pct":
                value1, value2 = 0.9645, 0.9753
            rows.append([section_header if offset == 0 else None, label, value1, value2, None])
            counter += 1
    rows += [
        ["FEDERAL INCOME TAX INFO", "Standard/Itemized Deductions", 13850.0, 14600.0, None],
        [None, "Bracket 1 Rate", 0.1, 0.1, None],
        [None, "Bracket 2 Rate", 0.12, 0.12, None],
        [None, "Bracket 1 Threshold", 0.0, 0.0, None],
        [None, "Bracket 2 Threshold", 11000.0, 11600.0, None],
        ["STATE INCOME TAX INFO", "Standard/Itemized Deductions", 5363.0, 5540.0, None],
        [None, "Exemption Credits", 144.0, 149.0, None],
        [None, "Bracket 1 Rate", 0.01, 0.01, None],
        [None, "Bracket 1 Threshold", 0.0, 0.0, None],
        ["MEDICARE TAX INFO", "Bracket 1 Rate", 0.0145, 0.0145, None],
        [None, "Bracket 1 Threshold", 0.0, 0.0, None],
        ["SOCIAL SECURITY TAX INFO", "Bracket 1 Rate", 0.062, 0.062, None],
        [None, "Bracket 1 Threshold", 0.0, 0.0, None],
        ["DISABILITY TAX INFO", "Bracket 1 Rate", 0.009, 0.01, None],
        [None, "Bracket 1 Threshold", 0.0, 0.0, None],
        ["CAPITAL GAINS TAX INFO", "Bracket 1 Rate", 0.0, 0.0, None],
        [None, "Bracket 1 Threshold", 0.0, 0.0, None],
        ["FEDERAL INCOME TAX", "Federal AGI", 99999.0, 99999.0, None],
        [None, "Taxes", 12345.0, 12345.0, None],
    ]
    return rows


def default_espp_rows() -> list[list]:
    # Columns B-E = modeler block; columns I-N = lots table. One row list covers both.
    def merged(row_bd: list, row_in: list) -> list:
        return [None] + row_bd + [None] * (8 - 1 - len(row_bd)) + row_in

    rows = [
        [None],
        merged(
            ["ESPP Modeler", "TOTAL CALENDAR YEAR", None, "<$25,000"],
            [
                "ESPP Purchase Date",
                "Qualifying Date",
                "Shares Purchased",
                "Subscription Price",
                "Purchase Date FMV",
                "Puchase Price",
            ],
        ),
        merged(
            [None, "FEBRUARY PURCHASE", "AUGUST PURCHASE", None],
            [datetime(2024, 2, 29), datetime(2025, 9, 1), 100.0, 40.0, 50.0, 34.0],
        ),
        merged(
            [None, "September-February Period", "March-August Period", None],
            [datetime(2024, 8, 30), datetime(2026, 2, 28), 90.0, 40.0, 55.0, 34.0],
        ),
        merged(
            ["Semi-Annual Base Salary", 50000.0, 60000.0, "enter your semi-annual base salary"],
            [datetime(2025, 2, 27), None, None, None, None, None],
        ),
        merged(
            ["Additional payments (i.e. bonuses)", 0.0, 100.0, None],
            [datetime(2025, 8, 29), None, None, None, None, None],
        ),
        merged(
            ["Total Eligible Earnings for 6-month Period", 50000.0, 60100.0, None],
            [datetime(2026, 2, 27), None, None, None, None, None],
        ),
        merged(
            ["ESPP Contribution Percentage", 0.1, 0.15, "enter your ESPP %"],
            [None, None, None, None, None, None],
        ),
    ]
    rows += [[None]] * 3
    rows += [
        [None, "ESPP Taxation Calculator"],
        [None, "Date of Sale", datetime(2025, 9, 1)],
    ]
    return rows


def default_paycheck_rows() -> list[list]:
    return [
        [None],
        [None, "Earnings", None, None, "Percentages", None],
        [None, "Annual Salary", 120000.0, None, "Traditional 401(k) %", 0.1],
        [None, "Gross Paycheck", 5000.0, None, "Roth 401(k) %", 0.0],
        [None, "Pretax Deductions", None, None, "AT 401(k) %", 0.02],
        [None, "Trad. 401(k)", 500.0, None, "Tax Withholding %", 0.250000000123456],
        [None, "Dental & Vision", 10.0, None, "ESPP %", 0.05],
        [None, "HSA", 50.0, None, None, None],
    ]


def default_focal_rows() -> list[list]:
    return [
        [None],
        [
            None,
            "Focal Year",
            "Current Base",
            "New Base",
            "Base Delta ($)",
            "Base Delta (%)",
            "Unvested RSUs",
            "Unvested Price",
            "Unvested Equity",
            "Refresh RSUs",
            "Grant Price",
            "Equity Delta ($)",
        ],
        [None, 2024.0, 110000.0, 120000.0, 10000, 0.09, 500.0, 89.66, 44830, 100.0, 90.0, 9000],
        [None, 2025.0, 120000.0, None, None, None, None, None, None, None, None, None],
        [None, 2026.0, None, None, None, None, None, None, None, None, None, None],
    ]


def default_reference_data_rows() -> list[list]:
    return [
        [
            "Symbol",
            "Name",
            "Sector",
            "Cost Per Share",
            "Last Price",
            "Dividend Yield",
            "Dividend per Share",
            "Payout Ratio",
            "Ex-Dividend Date",
        ],
        ["ACME", "Acme ETF", "ETF", 90.0, 100.5, "1.2", "2.5", 0, datetime(2024, 1, 31)],
        ["DIVC", "Div Corp", "Financials", "#REF!", "#N/A", "3.4", "1.25", 0, time(0, 0)],
        ["MUT1", "Mut Fund", "Mutual Fund", 25.0, 25.75, None, 0, 0, "N/A"],
    ]


def build_workbook(**overrides) -> bytes:
    """Build the synthetic workbook; override any sheet via keyword (rows list-of-lists)."""
    sheets = {
        "Paycheck Modeler": overrides.get("paycheck", default_paycheck_rows()),
        "ESPP": overrides.get("espp", default_espp_rows()),
        "Focal History": overrides.get("focal", default_focal_rows()),
        "Positions": overrides.get("positions", default_positions_rows()),
        "Spending": overrides.get("spending", default_spending_rows()),
        "Taxes": overrides.get("taxes", default_taxes_rows()),
        "Net Worth": overrides.get("net_worth", default_net_worth_rows()),
        "Portfolio": overrides.get("portfolio", default_portfolio_rows()),
        "ReferenceData": overrides.get("reference_data", default_reference_data_rows()),
    }
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for title, rows in sheets.items():
        if rows is None:
            continue  # simulate a missing sheet
        ws = wb.create_sheet(title=title)
        for row in rows:
            ws.append(row)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def load_readonly(data: bytes) -> openpyxl.Workbook:
    """Reload the built bytes exactly the way the service loads uploads."""
    return openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
