"""Importer orchestrator: workbook bytes -> parse all sheets -> apply or abort.

Same code path for dry-run and apply (spec section 5): the appliers always run; dry-run
rolls the session back instead of committing. Any parse error anywhere blocks the apply
entirely — the report still carries every sheet's errors and warnings.
"""

import io
import zipfile
from uuid import UUID, uuid4

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy.ext.asyncio import AsyncSession

from app.importer import apply as appliers
from app.importer import parsers
from app.importer.report import ImportReport
from app.models import ChangeLog, LifecycleRun
from app.services.snapshot import write_restore_point

PARSER_TABLE = (
    ("reference_data", "ReferenceData", parsers.parse_reference_data),
    ("positions", "Positions", parsers.parse_positions),
    ("portfolio", "Portfolio", parsers.parse_portfolio),
    ("net_worth", "Net Worth", parsers.parse_net_worth),
    ("spending", "Spending", parsers.parse_spending),
    ("taxes", "Taxes", parsers.parse_taxes),
    ("espp", "ESPP", parsers.parse_espp),
    ("paycheck", "Paycheck Modeler", parsers.parse_paycheck),
    ("focal_history", "Focal History", parsers.parse_focal_history),
)


class InvalidWorkbookError(ValueError):
    """The upload/file is not a readable .xlsx workbook."""


def _load_workbook(data: bytes):
    try:
        return openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except (zipfile.BadZipFile, InvalidFileException, KeyError, OSError) as exc:
        raise InvalidWorkbookError(str(exc)) from exc


def _import_counts(report: ImportReport) -> dict[str, dict[str, dict[str, int]]]:
    return {
        key: {
            entity: {"creates": c.creates, "updates": c.updates, "deletes": c.deletes}
            for entity, c in sheet.entities.items()
        }
        for key, sheet in report.sheets.items()
    }


def _summary_row(report: ImportReport, batch_id: UUID, actor: str | None) -> ChangeLog:
    """One op='batch' line for the whole apply (2026-09-03 data-lifecycle spec §9) — the
    per-row detail is the report's, stored on the run beside it."""
    counts = _import_counts(report)
    changes = sum(
        c.creates + c.updates + c.deletes
        for sheet in report.sheets.values()
        for c in sheet.entities.values()
    )
    return ChangeLog(
        batch_id=batch_id,
        source="import",
        actor=actor,
        label=f"Imported workbook — {changes} changes across {len(report.sheets)} sheets",
        table_name="*",
        pk={},
        op="batch",
        before=None,
        after={"sheets": counts},
        month=None,
    )


async def _record_run(
    db: AsyncSession, report: ImportReport, *, actor: str | None, batch_id: UUID | None
) -> None:
    """Every import — dry run, refused, applied — leaves a stored report (the card's report
    used to evaporate with React state). Its own commit: the apply's transaction is over."""
    db.add(
        LifecycleRun(
            kind="import_xlsx",
            dry_run=report.dry_run,
            ok=not report.has_errors,
            actor=actor,
            report=report.model_dump(mode="json"),
            batch_id=batch_id,
        )
    )
    await db.commit()


async def run_import(
    data: bytes, db: AsyncSession, *, dry_run: bool, actor: str | None = None
) -> ImportReport:
    report = ImportReport.new(dry_run=dry_run)
    workbook = _load_workbook(data)
    parsed: dict[str, object] = {}
    try:
        for key, sheet_name, parser in PARSER_TABLE:
            sheet_report = report.sheets[key]
            if sheet_name not in workbook.sheetnames:
                sheet_report.errors.append(f"sheet {sheet_name!r} not found in workbook")
                continue
            try:
                result = parser(workbook[sheet_name])
            except Exception as exc:
                # A structurally hollow sheet (e.g. zero rows) must be a row-context-free
                # sheet error, not a CLI traceback / HTTP 500 (Task 8 review finding).
                sheet_report.errors.append(f"{sheet_name}: parser crashed: {exc!r}")
                continue
            sheet_report.warnings.extend(result.issues.warnings)
            sheet_report.errors.extend(result.issues.errors)
            parsed[key] = result
    finally:
        workbook.close()
    if report.has_errors:
        await _record_run(db, report, actor=actor, batch_id=None)
        return report  # strict: errors anywhere block the whole apply (spec section 5)

    if not dry_run:
        # "This cannot be undone" leaves the import card (2026-09-03 data-lifecycle spec §9):
        # the current database is kept first, as its own committed run.
        await write_restore_point(db, actor=actor)
    batch_id: UUID | None = None
    try:
        by_name = await appliers.apply_reference_data(
            db, parsed["reference_data"], report.sheets["reference_data"]
        )
        await appliers.apply_positions(db, parsed["positions"], by_name, report.sheets["positions"])
        await appliers.apply_portfolio_history(db, parsed["portfolio"], report.sheets["portfolio"])
        await appliers.apply_net_worth(db, parsed["net_worth"], report.sheets["net_worth"])
        await appliers.apply_spending(db, parsed["spending"], report.sheets["spending"])
        await appliers.apply_taxes(db, parsed["taxes"], report.sheets["taxes"])
        await appliers.apply_espp(db, parsed["espp"], report.sheets["espp"])
        await appliers.apply_focal_history(
            db, parsed["focal_history"], report.sheets["focal_history"]
        )
        await appliers.apply_paycheck(
            db, parsed["paycheck"], parsed["focal_history"], report.sheets["paycheck"]
        )
        if report.has_errors or dry_run:  # apply_taxes can error on missing definitions
            await db.rollback()
        else:
            batch_id = uuid4()
            db.add(_summary_row(report, batch_id, actor))  # rides the apply's own commit
            await db.commit()
            report.applied = True
    except Exception:
        await db.rollback()
        raise
    await _record_run(db, report, actor=actor, batch_id=batch_id)
    return report
